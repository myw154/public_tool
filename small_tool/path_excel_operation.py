# 读取Excel的几种方式
import os

from openpyxl import load_workbook, Workbook
def del_excel(file_path):
    new_file = Workbook()
    new_file.remove_sheet(new_file.active)  # 删除默认的活动的表格
    new_table_name = ''
    new_tab = self.new_file.create_sheet(one_tab_name)  # 创建一个新的表格

    open_file = load_workbook(file_path)
    table_name_list = open_file.sheetnames
    for one_table_name in talble_name_list:
        tab_obj = open_file[one_table_name]
        max_row = tab_obj.max_row
        max_col = tab_obj.max_column
        for i in range(1, max_row+1):
            for j in range(1, max_col):
                one_value = tab_obj.cell(i, j).value
    new_file.save('new_file_name')

import csv
def read_csv(file_path):
    with open('example.csv', 'r') as file:
        reader = csv.reader(file)
        for row in reader:
            print(row)


import pandas as pd
def pandas_read(file_path, save_path):
    # 方法一： 使用pd.read_excel
    # df1=pd.read_excel(file_path,sheet_name=None)
    # print(list(df1.keys()))
    # for i in df1.keys():
        # print(df1[i])

    # 方法二： 使用pd.ExcelFile
    df2 = pd.ExcelFile(file_path)
    sheet_name_list = df2.sheet_names
    for name in sheet_name_list:
        df_pre = df2.parse(sheet_name=name)
        for ondex, one_ in df_pre.iteritems():
            print(index, one_)
        print(df_pre)
    df2.close()
    df2.save(save_path)


